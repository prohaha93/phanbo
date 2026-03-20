import pandas as pd
import numpy as np
import pulp
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO          # ADDED: để tạo luồng nhị phân

# ============================================================
# COLOR PALETTE (giữ nguyên)
# ============================================================
C_DARK_BLUE   = "FF1F4E79"
C_MID_BLUE    = "FF2E75B6"
C_LIGHT_BLUE  = "FF9DC3E6"
C_PALE_BLUE   = "FFD6E4F0"
C_ALT_ROW     = "FFEBF3FB"
C_WHITE       = "FFFFFFFF"
C_YELLOW      = "FFFFF2CC"
C_GREEN       = "FF375623"
C_TITLE_BG    = "FFDEEAF1"
C_TITLE_BG_M  = "FFD6E4F0"
C_ORANGE_FILL = "FFFCE4D6"
C_ORANGE_FONT = "FF833C00"
C_GREY_FONT   = "FFBFBFBF"
C_HEADER_BG   = "FFDEEAF1"

FONT_NAME = "Calibri"

def _font(bold=False, color="FF000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(border_style="medium", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _style(ws, coord, value=None, bold=False, font_color="FF000000",
           fill_color=None, align="center", wrap=False, border=True, size=10):
    cell = ws[coord] if isinstance(coord, str) else coord
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, color=font_color, size=size)
    if fill_color:
        cell.fill = _fill(fill_color)
    cell.alignment = _align(h=align, wrap=wrap)
    if border:
        cell.border = _thin_border()
    return cell

# ============================================================
# HÀM CHÍNH: run_optimization
# Nhận đầu vào là file Excel (đường dẫn hoặc file-like object)
# Trả về: (excel_bytes, total_rows, objective_value)
# ============================================================
def run_optimization(input_file):
    """
    Thực hiện toàn bộ quá trình đọc dữ liệu, tối ưu và tạo file Excel kết quả.
    input_file: đường dẫn (string) hoặc đối tượng file-like (BytesIO) chứa file Excel đầu vào.
    Trả về: (excel_bytes, total_rows, objective_value)
    """
    # ============================================================
    # 1. Read and parse original data (sử dụng input_file thay vì tên cứng)
    # ============================================================
    # MODIFIED: dùng pd.ExcelFile trên input_file
    xls = pd.ExcelFile(input_file)

    # --- Sheet 1: MOVEHOUR-WEIGHTCLASS → demand per (hour, STS, bay, wc) ---
    df1 = pd.read_excel(xls, sheet_name='MOVEHOUR-WEIGHTCLASS', header=None)
    sts_bay_map = {}
    for col in range(2, df1.shape[1]):
        sts = df1.iloc[0, col]
        bay = df1.iloc[1, col]
        if pd.notna(sts) and pd.notna(bay):
            sts_bay_map[col] = (str(sts).strip(), str(bay).strip())

    demands = {}
    current_hour = None
    for idx in range(2, df1.shape[0]):
        row = df1.iloc[idx]
        hour = row[0]
        if pd.isna(hour):
            hour = current_hour
        else:
            current_hour = hour
        weight = row[1]
        if pd.isna(weight):
            continue
        weight = int(weight)
        for col in range(2, df1.shape[1]):
            qty = row[col]
            if pd.notna(qty) and qty != '':
                qty = int(qty)
                if qty > 0:
                    sts, bay = sts_bay_map[col]
                    key = (hour, sts, bay)
                    if key not in demands:
                        demands[key] = {}
                    demands[key][weight] = demands[key].get(weight, 0) + qty

    job_keys = list(demands.keys())

    all_hours_sorted = sorted(set(h for (h, s, b) in job_keys))
    hour_rank = {h: i for i, h in enumerate(all_hours_sorted)}

    jobs_by_hour = {}
    for (h, s, b) in job_keys:
        jobs_by_hour.setdefault(h, []).append((s, b))

    # --- Sheet 2: BLOCK-WEIGHT CLASS → aggregate supply per (block, wc) ---
    df2 = pd.read_excel(xls, sheet_name='BLOCK-WEIGHT CLASS', header=0)
    supply = {}
    for idx, row in df2.iterrows():
        block = str(row.iloc[0]).strip()
        if block == 'nan':
            continue
        w1 = int(row.iloc[1]) if pd.notna(row.iloc[1]) else 0
        w2 = int(row.iloc[2]) if pd.notna(row.iloc[2]) else 0
        w3 = int(row.iloc[3]) if pd.notna(row.iloc[3]) else 0
        w4 = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
        w5 = int(row.iloc[5]) if len(row) > 5 and pd.notna(row.iloc[5]) else 0
        supply[block] = {1: w1, 2: w2, 3: w3, 4: w4, 5: w5}

    weight_classes = [1, 2, 3, 4, 5]
    blocks = [b for b in supply if any(supply[b][w] > 0 for w in weight_classes)]

    # --- Sheet 3 (DATA file): container-level layout ---
    container_data_available = False
    try:
        df_containers = pd.read_excel(xls, sheet_name='DATA', header=0)
        required = {'YARD', 'YB', 'YR', 'YT', 'Unnamed: 1', 'Unnamed: 2'}
        if required.issubset(df_containers.columns):
            df_containers = df_containers.dropna(subset=['Unnamed: 1', 'Unnamed: 2', 'YB', 'YR', 'YT']).copy()
            df_containers['REAL_WC'] = df_containers['Unnamed: 1'].astype(float).astype(int)
            df_containers['YARD_POS'] = df_containers['Unnamed: 2'].astype(str).str.strip()
            if 'Unnamed: 3' in df_containers.columns:
                df_containers['REAL_CONT_ID'] = df_containers['Unnamed: 3'].fillna('').astype(str).str.strip()
            else:
                df_containers['REAL_CONT_ID'] = ''
            df_containers['CONT_ID'] = df_containers['YARD_POS']
            df_containers['YARD'] = df_containers['YARD'].astype(str).str.strip()
            df_containers['YB'] = df_containers['YB'].astype(float).astype(int)
            df_containers['YR'] = df_containers['YR'].astype(float).astype(int)
            df_containers['YT'] = df_containers['YT'].astype(float).astype(int)
            container_data_available = True
            print("Container-level DATA sheet found – stacking rules will be applied.")
            print(f"  {len(df_containers)} containers loaded with real WC from col B.")
        else:
            print("DATA sheet missing required columns – stacking rules skipped.")
    except Exception as e:
        print(f"No DATA sheet found – stacking rules skipped. ({e})")

    # ============================================================
    # 1b. Build container-level stacking structures (giữ nguyên)
    # ============================================================
    yb_wc_supply   = {}
    stack_ordering = {}
    blocking_pairs = []

    if container_data_available:
        df_c = df_containers[['YARD','YB','YR','YT','REAL_WC','YARD_POS','REAL_CONT_ID']].copy()
        for block in blocks:
            block_df = df_c[df_c['YARD'] == block]
            if block_df.empty:
                continue
            yb_wc_supply[block] = {}
            stack_ordering[block] = {}
            for yb, yb_df in block_df.groupby('YB'):
                yb_wc_supply[block][yb] = {}
                stack_ordering[block][yb] = {}
                for wc, cnt in yb_df.groupby('REAL_WC').size().items():
                    yb_wc_supply[block][yb][wc] = int(cnt)
                for yr, yr_df in yb_df.groupby('YR'):
                    ordered = yr_df.sort_values('YT', ascending=False)[['YT','REAL_WC']].values.tolist()
                    stack_ordering[block][yb][yr] = [(int(t), int(w)) for t, w in ordered]
                for yr, tiers in stack_ordering[block][yb].items():
                    wcs_above = []
                    for tier, wc in tiers:
                        if wcs_above:
                            for (prev_wc, prev_tier) in wcs_above:
                                if prev_wc != wc:
                                    blocking_pairs.append((block, yb, yr, prev_tier, prev_wc, tier, wc))
                        wcs_above.append((wc, tier))
        print(f"Stacking structures built: {len(blocking_pairs)} cross-WC blocking pairs found.")

    # ============================================================
    # 2. Check total demand vs supply per weight class
    # ============================================================
    total_demand = {w: 0 for w in weight_classes}
    for job in job_keys:
        for w, qty in demands[job].items():
            total_demand[w] += qty
    total_supply = {w: 0 for w in weight_classes}
    for block in blocks:
        for w in weight_classes:
            total_supply[w] += supply[block][w]

    print("Total demand per weight class:")
    for w in weight_classes:
        print(f"  Class {w}: {total_demand[w]}")
    print("Total supply per weight class:")
    for w in weight_classes:
        print(f"  Class {w}: {total_supply[w]}")

    for w in weight_classes:
        if total_demand[w] != total_supply[w]:
            print(f"ERROR: Mismatch for weight class {w}: demand={total_demand[w]}, supply={total_supply[w]}")
            # MODIFIED: thay exit(1) bằng raise ValueError
            raise ValueError(f"Tổng cầu và cung không khớp cho weight class {w}")

    # ============================================================
    # 3. Build and solve the optimisation model
    # ============================================================
    prob = pulp.LpProblem("Minimize_Clashes_With_Stacking", pulp.LpMinimize)

    y_vars = {}
    for (h, s, bay) in job_keys:
        for b in blocks:
            y_vars[(h, s, bay, b)] = pulp.LpVariable(f"y_{h}_{s}_{bay}_{b}", cat='Binary')

    x_vars = {}
    for (h, s, bay) in job_keys:
        for w in demands[(h, s, bay)]:
            for b in blocks:
                x_vars[(h, s, bay, b, w)] = pulp.LpVariable(
                    f"x_{h}_{s}_{bay}_{b}_{w}", lowBound=0, cat='Integer')

    u_vars = {}
    e_vars = {}
    for h in jobs_by_hour:
        for b in blocks:
            u_vars[(h, b)] = pulp.LpVariable(f"u_{h}_{b}", lowBound=0, cat='Integer')
            e_vars[(h, b)] = pulp.LpVariable(f"e_{h}_{b}", lowBound=0, cat='Integer')
            prob += u_vars[(h, b)] == pulp.lpSum(
                y_vars[(h, s, bay, b)] for (s, bay) in jobs_by_hour[h])
            prob += e_vars[(h, b)] >= u_vars[(h, b)] - 1

    prob += pulp.lpSum(e_vars.values())

    # C1. Demand satisfaction
    for (h, s, bay) in job_keys:
        for w, d in demands[(h, s, bay)].items():
            prob += pulp.lpSum(x_vars[(h, s, bay, b, w)] for b in blocks) == d

    # C2. Aggregate supply cap per (block, wc)
    for b in blocks:
        for w in weight_classes:
            prob += pulp.lpSum(
                x_vars[(h, s, bay, b, w)]
                for (h, s, bay) in job_keys if w in demands[(h, s, bay)]
            ) <= supply[b][w]

    # C3. Linking x -> y
    for (h, s, bay) in job_keys:
        for w in demands[(h, s, bay)]:
            d = demands[(h, s, bay)][w]
            for b in blocks:
                prob += x_vars[(h, s, bay, b, w)] <= d * y_vars[(h, s, bay, b)]

    # Solve
    solver = pulp.PULP_CBC_CMD(msg=True, timeLimit=300)
    prob.solve(solver)

    status = prob.status
    print(f"Status: {pulp.LpStatus[status]}")
    if status == pulp.LpStatusInfeasible:
        print("Model is infeasible. Check supply/demand totals.")
        raise ValueError("Không tìm được lời giải do dữ liệu không khả thi.")
    elif status not in (1,):
        print("No optimal solution found within time limit – using best solution found.")

    # ============================================================
    # 4. Extract result + map individual containers (giữ nguyên logic)
    # ============================================================
    result_rows = []
    for (h, s, bay, b) in y_vars:
        if pulp.value(y_vars[(h, s, bay, b)]) is not None and \
           pulp.value(y_vars[(h, s, bay, b)]) > 0.5:
            for w in demands[(h, s, bay)]:
                qty = pulp.value(x_vars[(h, s, bay, b, w)])
                if qty is not None and qty > 0.5:
                    result_rows.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'QUANTITIES': int(round(qty))
                    })

    df_result = pd.DataFrame(result_rows)
    df_result.sort_values(['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK'], inplace=True)

    # ------------------------------------------------------------------
    # 4b. Map individual containers (giữ nguyên)
    # ------------------------------------------------------------------
    df_result_detail = []
    if container_data_available:
        # Build container pool
        pool = {}
        for _, row in df_containers[['YARD','YB','YR','YT','REAL_WC','YARD_POS','REAL_CONT_ID']].iterrows():
            blk = row['YARD']
            pool.setdefault(blk, []).append({
                'yb': int(row['YB']), 'yr': int(row['YR']), 'yt': int(row['YT']),
                'wc': int(row['REAL_WC']),
                'yard_pos': row['YARD_POS'],
                'real_cont_id': row['REAL_CONT_ID'],
                'picked': False, 'pick_h': None
            })

        def accessible_at(cont, containers, h_rank_val):
            for c in containers:
                if c is cont:
                    continue
                if c['yb'] == cont['yb'] and c['yr'] == cont['yr'] and c['yt'] > cont['yt']:
                    if not c['picked']:
                        return False
                    if c['pick_h'] is not None and hour_rank[c['pick_h']] > h_rank_val:
                        return False
            return True

        def pick_n(block, wc, qty, h, s_job, bay_job, h_rank_val, result_list):
            containers = pool[block]
            remaining = qty
            while remaining > 0:
                cands = [c for c in containers
                         if not c['picked'] and c['wc'] == wc
                         and accessible_at(c, containers, h_rank_val)]
                if not cands:
                    break
                yb_cnt = {}
                for c in cands:
                    yb_cnt[c['yb']] = yb_cnt.get(c['yb'], 0) + 1
                cands.sort(key=lambda c: (
                    -yb_cnt[c['yb']],
                    c['yb'],
                    c['yr'],
                    -c['yt']
                ))
                best = cands[0]
                best['picked'] = True
                best['pick_h'] = h
                result_list.append({
                    'MOVE HOUR': h,
                    'CONTAINER ID': best['real_cont_id'],
                    'STS': s_job, 'BAY': bay_job,
                    'ASSIGNED BLOCK': block, 'WEIGHT CLASS': wc,
                    'QUANTITIES': qty,
                    'YB': best['yb'], 'YR': best['yr'], 'YT': best['yt'],
                    'YARD POSITION': best['yard_pos']
                })
                remaining -= 1
            return remaining

        df_result_sorted = df_result.copy()
        df_result_sorted['_hr'] = df_result_sorted['MOVE HOUR'].map(hour_rank)
        df_result_sorted.sort_values(['_hr','STS','BAY','ASSIGNED BLOCK','WEIGHT CLASS'],
                                      inplace=True)

        deferred = []
        for h in all_hours_sorted:
            h_rank_val = hour_rank[h]
            hour_asgns = df_result_sorted[df_result_sorted['MOVE HOUR'] == h]
            for _, asg in hour_asgns.iterrows():
                s, bay_job, b = asg['STS'], asg['BAY'], asg['ASSIGNED BLOCK']
                w, qty = int(asg['WEIGHT CLASS']), int(asg['QUANTITIES'])
                if b not in pool:
                    df_result_detail.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay_job,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'CONTAINER ID': '', 'QUANTITIES': qty,
                        'YB': '', 'YR': '', 'YT': '', 'YARD POSITION': ''
                    })
                    continue
                rem = pick_n(b, w, qty, h, s, bay_job, h_rank_val, df_result_detail)
                if rem > 0:
                    deferred.append({'b': b, 'wc': w, 'qty': rem,
                                      'h_orig': h, 's': s, 'bay': bay_job,
                                      'h_rank_min': h_rank_val})
            still_deferred = []
            for d in deferred:
                rem = pick_n(d['b'], d['wc'], d['qty'], h,
                             d['s'], d['bay'], h_rank_val, df_result_detail)
                if rem > 0:
                    d2 = d.copy()
                    d2['qty'] = rem
                    still_deferred.append(d2)
            deferred = still_deferred

        rehandle_count = sum(d['qty'] for d in deferred)
        if rehandle_count > 0:
            print(f"  INFO: {rehandle_count} containers require re-handling "
                  f"(cross-WC stacking unavoidable):")
            for d in deferred:
                print(f"    Block {d['b']} WC{d['wc']} x{d['qty']} "
                      f"(originally at {d['h_orig']})")
        else:
            print("  All containers assigned with no re-handling required.")

        df_result_detail = pd.DataFrame(df_result_detail)
    else:
        df_result_detail = df_result.copy()
        df_result_detail.insert(1, 'CONTAINER ID', '')
        df_result_detail['YB'] = ''
        df_result_detail['YR'] = ''
        df_result_detail['YT'] = ''
        df_result_detail['YARD POSITION'] = ''

    df_result_detail.sort_values(
        ['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK', 'WEIGHT CLASS', 'YB', 'YR', 'YT'],
        inplace=True
    )

    # ============================================================
    # 5. Prepare MATRIX data (giữ nguyên)
    # ============================================================
    df_matrix_base = df_result.groupby(
        ['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK'], as_index=False
    )['QUANTITIES'].sum()

    sts_list = sorted(df_result['STS'].unique(), key=lambda x: int(x.replace('STS', '')))
    sts_bay_blocks = {}

    def _first_hour(sts, bay, df):
        hours = df[(df['STS'] == sts) & (df['BAY'] == bay)]['MOVE HOUR'].unique()
        return sorted(hours)[0]

    for sts in sts_list:
        bays = df_result[df_result['STS'] == sts]['BAY'].unique()
        bays = sorted(bays, key=lambda bay: _first_hour(sts, bay, df_result))
        sts_bay_blocks[sts] = {}
        for bay in bays:
            blks = df_result[(df_result['STS'] == sts) & (df_result['BAY'] == bay)]['ASSIGNED BLOCK'].unique()
            sts_bay_blocks[sts][bay] = sorted(blks)

    matrix_cols = []
    for sts in sts_list:
        for bay in sts_bay_blocks[sts]:
            for block in sts_bay_blocks[sts][bay]:
                matrix_cols.append((sts, bay, block))

    hour_list = sorted(df_matrix_base['MOVE HOUR'].unique())
    matrix_data = {}
    for h in hour_list:
        matrix_data[h] = {}
        for col in matrix_cols:
            matrix_data[h][col] = 0

    for _, row in df_matrix_base.iterrows():
        key = (row['STS'], row['BAY'], row['ASSIGNED BLOCK'])
        if key in matrix_data.get(row['MOVE HOUR'], {}):
            matrix_data[row['MOVE HOUR']][key] = row['QUANTITIES']

    # ============================================================
    # 6. Prepare DETAIL groups (giữ nguyên)
    # ============================================================
    detail_groups_by_sts = {}
    for sts in sts_list:
        detail_groups_by_sts[sts] = []
        for bay in sts_bay_blocks[sts]:
            df_sub = df_result[(df_result['STS'] == sts) & (df_result['BAY'] == bay)]
            rows_idx = df_sub[['MOVE HOUR', 'WEIGHT CLASS']].drop_duplicates().sort_values(['MOVE HOUR', 'WEIGHT CLASS'])
            blks = sorted(df_sub['ASSIGNED BLOCK'].unique())
            pivot = df_sub.pivot_table(index=['MOVE HOUR', 'WEIGHT CLASS'], columns='ASSIGNED BLOCK',
                                       values='QUANTITIES', fill_value=0, aggfunc='sum')
            pivot = pivot.reindex(
                index=pd.MultiIndex.from_tuples([(h, w) for h, w in rows_idx.values]),
                columns=blks, fill_value=0
            )
            table_rows = []
            for (hour, wc) in pivot.index:
                row_data = [hour, wc] + [int(pivot.loc[(hour, wc), b]) for b in blks] + [int(pivot.loc[(hour, wc)].sum())]
                table_rows.append(row_data)

            col_totals = [None, None]
            for b in blks:
                col_totals.append(int(pivot[b].sum()))
            col_totals.append(int(pivot.values.sum()))

            detail_groups_by_sts[sts].append({
                'bay': bay,
                'blocks': blks,
                'rows': table_rows,
                'col_totals': col_totals,
                'num_rows': len(table_rows)
            })

    sts_table_widths = {}
    for sts in sts_list:
        max_w = max(2 + len(g['blocks']) + 1 for g in detail_groups_by_sts[sts])
        sts_table_widths[sts] = max_w

    # ============================================================
    # 7. Write Excel with openpyxl
    # ============================================================
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- SHEET: MOVEHOUR-WEIGHTCLASS ---
    ws_mh = wb.create_sheet('MOVEHOUR-WEIGHTCLASS')
    for r_idx, row in enumerate(df1.values, 1):
        for c_idx, val in enumerate(row, 1):
            ws_mh.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)

    # --- SHEET: BLOCK-WEIGHT CLASS ---
    ws_bw = wb.create_sheet('BLOCK-WEIGHT CLASS')
    headers = list(df2.columns)
    for c_idx, h in enumerate(headers, 1):
        ws_bw.cell(row=1, column=c_idx, value=h)
    for r_idx, row in enumerate(df2.values, 2):
        for c_idx, val in enumerate(row, 1):
            ws_bw.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)

    # ============================================================
    # SHEET: RESULT (với CONT LIST và CONTAINER ID)
    # ============================================================
    ws_result = wb.create_sheet('RESULT')

    # Pre-compute CONT LIST
    cont_list_map = {}
    if container_data_available and 'CONTAINER ID' in df_result_detail.columns:
        for (mh, bay), grp in df_result_detail.groupby(['MOVE HOUR', 'BAY']):
            ids = [str(v).strip() for v in grp['CONTAINER ID'] if str(v).strip() not in ('', 'nan')]
            cont_list_map[(mh, bay)] = ', '.join(ids) if ids else ''

    core_cols     = ['MOVE HOUR', 'CONT LIST', 'CONTAINER ID', 'STS', 'BAY',
                      'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']
    position_cols = ['YB', 'YR', 'YT', 'YARD POSITION']

    if container_data_available:
        all_result_cols = core_cols + position_cols
    else:
        all_result_cols = ['MOVE HOUR', 'STS', 'BAY',
                            'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']

    CONT_LIST_COLS = {'CONT LIST'}
    CONT_ID_COLS   = {'CONTAINER ID'}
    POSITION_COLS  = set(position_cols)

    # Header row
    for c_idx, cn in enumerate(all_result_cols, 1):
        cell = ws_result.cell(row=1, column=c_idx, value=cn)
        if cn in CONT_LIST_COLS:
            cell.fill = _fill(C_PALE_BLUE)
        elif cn in CONT_ID_COLS:
            cell.fill = _fill(C_MID_BLUE)
        elif cn in POSITION_COLS:
            cell.fill = _fill(C_LIGHT_BLUE)
        else:
            cell.fill = _fill(C_DARK_BLUE)
        cell.font      = _font(bold=True, color=C_WHITE)
        cell.alignment = _align(wrap=True)
        cell.border    = _thin_border()

    df_rd = df_result_detail.reset_index(drop=True)
    n_rows = len(df_rd)

    # Track merge for CONT LIST
    cont_list_col_idx = all_result_cols.index('CONT LIST') + 1 if container_data_available else None
    merge_groups = []
    if container_data_available:
        prev_key = None
        grp_start = 2
        for i, (_, row) in enumerate(df_rd.iterrows()):
            cur_key = (row.get('MOVE HOUR', ''), row.get('BAY', ''))
            excel_row = i + 2
            if cur_key != prev_key:
                if prev_key is not None:
                    merge_groups.append((prev_key, grp_start, excel_row - 1,
                                         cont_list_map.get(prev_key, '')))
                prev_key = cur_key
                grp_start = excel_row
        if prev_key is not None:
            merge_groups.append((prev_key, grp_start, n_rows + 1,
                                 cont_list_map.get(prev_key, '')))

    # Write data rows
    group_key   = None
    group_shade = C_ALT_ROW
    for r_idx, (_, row) in enumerate(df_rd.iterrows(), 2):
        this_key = (row.get('MOVE HOUR'), row.get('STS'), row.get('BAY'),
                    row.get('ASSIGNED BLOCK'), row.get('WEIGHT CLASS'))
        if this_key != group_key:
            group_shade = C_WHITE if group_shade == C_ALT_ROW else C_ALT_ROW
            group_key = this_key

        for c_idx, cn in enumerate(all_result_cols, 1):
            if cn == 'CONT LIST':
                val = None
            else:
                val = row.get(cn, '')
                if cn in ('YB', 'YR', 'YT') and val != '':
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        pass
                if val == '' or (isinstance(val, float) and str(val) == 'nan'):
                    val = None
            cell = ws_result.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = _font(color='FF000000')
            cell.fill      = _fill(group_shade)
            cell.alignment = _align(wrap=(cn == 'CONT LIST'))
            cell.border    = _thin_border()

    # Merge CONT LIST cells
    if container_data_available:
        for (mh, bay), r_start, r_end, list_text in merge_groups:
            cell = ws_result.cell(row=r_start, column=cont_list_col_idx, value=list_text or None)
            cell.font      = _font(color='FF000000', size=9)
            cell.fill      = _fill(C_PALE_BLUE)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border    = _thin_border()
            if r_end > r_start:
                ws_result.merge_cells(
                    start_row=r_start, start_column=cont_list_col_idx,
                    end_row=r_end,     end_column=cont_list_col_idx
                )
                ws_result.cell(row=r_start, column=cont_list_col_idx).alignment = \
                    Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Row heights
        for (mh, bay), r_start, r_end, list_text in merge_groups:
            span = r_end - r_start + 1
            n_ids = len([x for x in list_text.split(',') if x.strip()]) if list_text else 0
            rows_needed = max(1, -(-n_ids // max(1, span)))
            h = max(15, min(60, rows_needed * 13))
            for r in range(r_start, r_end + 1):
                ws_result.row_dimensions[r].height = h

    # Column widths
    col_widths = {
        'MOVE HOUR': 14,      'CONT LIST': 45,      'CONTAINER ID': 20,
        'STS': 10,            'BAY': 10,
        'ASSIGNED BLOCK': 16, 'WEIGHT CLASS': 14,   'QUANTITIES': 12,
        'YB': 8,              'YR': 8,              'YT': 8,
        'YARD POSITION': 18,
    }
    for c_idx, cn in enumerate(all_result_cols, 1):
        ws_result.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(cn, 14)

    # ============================================================
    # SHEET: MATRIX (giữ nguyên)
    # ============================================================
    ws_matrix = wb.create_sheet('MATRIX')
    total_matrix_cols = len(matrix_cols) + 2
    title_cell = ws_matrix.cell(row=1, column=1, value='MA TRẬN PHÂN BỔ BLOCK  ▸  MOVE HOUR × STS / BAY / BLOCK')
    title_cell.font = _font(bold=True, color=C_DARK_BLUE, size=11)
    title_cell.fill = _fill(C_TITLE_BG_M)
    title_cell.alignment = _align()
    title_cell.border = _thin_border()
    ws_matrix.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_matrix_cols)
    ws_matrix.row_dimensions[1].height = 16.8

    # Merge A2:A4 for "MOVE\nHOUR"
    mh_cell = ws_matrix.cell(row=2, column=1, value='MOVE\nHOUR')
    mh_cell.font = _font(bold=True, color=C_WHITE)
    mh_cell.fill = _fill(C_DARK_BLUE)
    mh_cell.alignment = _align(wrap=True)
    mh_cell.border = _thin_border()
    ws_matrix.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)

    total_col = total_matrix_cols
    tc = ws_matrix.cell(row=2, column=total_col, value='TOTAL')
    tc.font = _font(bold=True, color=C_WHITE)
    tc.fill = _fill(C_GREEN)
    tc.alignment = _align()
    tc.border = _thin_border()
    ws_matrix.merge_cells(start_row=2, start_column=total_col, end_row=4, end_column=total_col)

    col_offset = 2
    for sts in sts_list:
        sts_start = col_offset
        for bay in sts_bay_blocks[sts]:
            bay_start = col_offset
            for block in sts_bay_blocks[sts][bay]:
                bc = ws_matrix.cell(row=4, column=col_offset, value=block)
                bc.font = _font(bold=True, color=C_DARK_BLUE)
                bc.fill = _fill(C_LIGHT_BLUE)
                bc.alignment = _align()
                bc.border = _thin_border()
                col_offset += 1
            bay_end = col_offset - 1
            bayc = ws_matrix.cell(row=3, column=bay_start, value=bay)
            bayc.font = _font(bold=True, color=C_WHITE)
            bayc.fill = _fill(C_MID_BLUE)
            bayc.alignment = _align()
            bayc.border = _thin_border()
            if bay_start < bay_end:
                ws_matrix.merge_cells(start_row=3, start_column=bay_start, end_row=3, end_column=bay_end)
                for mc in range(bay_start+1, bay_end+1):
                    ws_matrix.cell(row=3, column=mc).border = _thin_border()
            for mc in range(bay_start, bay_end+1):
                ws_matrix.cell(row=2, column=mc).border = _thin_border()
        sts_end = col_offset - 1
        stsc = ws_matrix.cell(row=2, column=sts_start, value=sts)
        stsc.font = _font(bold=True, color=C_WHITE)
        stsc.fill = _fill(C_DARK_BLUE)
        stsc.alignment = _align()
        stsc.border = _thin_border()
        if sts_start < sts_end:
            ws_matrix.merge_cells(start_row=2, start_column=sts_start, end_row=2, end_column=sts_end)

    for r_idx, hour in enumerate(hour_list):
        excel_row = 5 + r_idx
        fill_color = C_ALT_ROW if (r_idx % 2 == 0) else C_WHITE
        hc = ws_matrix.cell(row=excel_row, column=1, value=hour)
        hc.font = _font(bold=True, color=C_DARK_BLUE)
        hc.fill = _fill(C_PALE_BLUE)
        hc.alignment = _align()
        hc.border = _thin_border()
        row_total = 0
        for c_idx, col_key in enumerate(matrix_cols, 2):
            val = matrix_data[hour].get(col_key, 0)
            dc = ws_matrix.cell(row=excel_row, column=c_idx)
            if val == 0:
                dc.value = '—'
                dc.font = _font(color=C_GREY_FONT)
            else:
                dc.value = val
                dc.font = _font(color="FF000000")
                row_total += val
            dc.fill = _fill(fill_color)
            dc.alignment = _align()
            dc.border = _thin_border()
        rtc = ws_matrix.cell(row=excel_row, column=total_col, value=row_total)
        rtc.font = _font(bold=True, color="FF000000")
        rtc.fill = _fill(C_YELLOW)
        rtc.alignment = _align()
        rtc.border = _thin_border()

    total_row = 5 + len(hour_list)
    trc = ws_matrix.cell(row=total_row, column=1, value='TOTAL')
    trc.font = _font(bold=True, color=C_WHITE)
    trc.fill = _fill(C_GREEN)
    trc.alignment = _align()
    trc.border = _thin_border()

    grand_total = 0
    for c_idx, col_key in enumerate(matrix_cols, 2):
        col_sum = sum(matrix_data[h].get(col_key, 0) for h in hour_list)
        tc2 = ws_matrix.cell(row=total_row, column=c_idx, value=col_sum)
        tc2.font = _font(bold=True, color="FF000000")
        tc2.fill = _fill(C_YELLOW)
        tc2.alignment = _align()
        tc2.border = _thin_border()
        grand_total += col_sum

    gtc = ws_matrix.cell(row=total_row, column=total_col, value=grand_total)
    gtc.font = _font(bold=True, color=C_WHITE)
    gtc.fill = _fill(C_GREEN)
    gtc.alignment = _align()
    gtc.border = _thin_border()

    ws_matrix.column_dimensions['A'].width = 12
    for c in range(2, total_matrix_cols + 1):
        ws_matrix.column_dimensions[get_column_letter(c)].width = 8

    # ============================================================
    # SHEET: DETAIL (giữ nguyên)
    # ============================================================
    ws_detail = wb.create_sheet('DETAIL')
    sts_col_start = {}
    current_col = 1
    for sts in sts_list:
        sts_col_start[sts] = current_col
        current_col += sts_table_widths[sts] + 1
    total_detail_cols = current_col - 2

    title_d = ws_detail.cell(row=1, column=1, value='TỔNG HỢP CHI TIẾT  ▸  STS / BAY / MOVE HOUR / WC / BLOCK')
    title_d.font = _font(bold=True, color=C_DARK_BLUE, size=11)
    title_d.fill = _fill(C_HEADER_BG)
    title_d.alignment = _align()
    title_d.border = _thin_border()
    ws_detail.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_detail_cols)

    def write_detail_sts(ws, sts, start_col, table_width, groups, row_start):
        end_col = start_col + table_width - 1
        current_row = row_start
        # STS header
        sts_cell = ws.cell(row=current_row, column=start_col, value=sts)
        sts_cell.font = _font(bold=True, color=C_WHITE)
        sts_cell.fill = _fill(C_DARK_BLUE)
        sts_cell.alignment = _align()
        sts_cell.border = _thin_border()
        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
        for mc in range(start_col+1, end_col+1):
            ws.cell(row=current_row, column=mc).border = _thin_border()
        current_row += 1

        for group in groups:
            bay = group['bay']
            blks = group['blocks']
            tbl_rows = group['rows']
            col_totals = group['col_totals']

            # BAY header
            bay_cell = ws.cell(row=current_row, column=start_col, value=bay)
            bay_cell.font = _font(bold=True, color=C_WHITE)
            bay_cell.fill = _fill(C_MID_BLUE)
            bay_cell.alignment = _align()
            bay_cell.border = _thin_border()
            ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
            for mc in range(start_col+1, end_col+1):
                ws.cell(row=current_row, column=mc).border = _thin_border()
            current_row += 1

            # Column headers
            col = start_col
            for hdr in ['MOVE HOUR', 'WC'] + blks:
                hc = ws.cell(row=current_row, column=col, value=hdr)
                hc.font = _font(bold=True, color=C_WHITE)
                hc.fill = _fill(C_DARK_BLUE)
                hc.alignment = _align()
                hc.border = _thin_border()
                col += 1
            while col <= end_col - 1:
                pc = ws.cell(row=current_row, column=col)
                pc.font = _font(bold=True, color=C_WHITE)
                pc.fill = _fill(C_DARK_BLUE)
                pc.alignment = _align()
                pc.border = _thin_border()
                col += 1
            tc_hdr = ws.cell(row=current_row, column=end_col, value='TOTAL')
            tc_hdr.font = _font(bold=True, color=C_WHITE)
            tc_hdr.fill = _fill(C_GREEN)
            tc_hdr.alignment = _align()
            tc_hdr.border = _thin_border()
            current_row += 1

            # Data rows
            hour_color_map = {}
            color_toggle = True
            for rd in tbl_rows:
                h_val = rd[0]
                if h_val not in hour_color_map:
                    hour_color_map[h_val] = C_ALT_ROW if color_toggle else C_WHITE
                    color_toggle = not color_toggle

            prev_hour = None
            for rd in tbl_rows:
                h_val = rd[0]
                wc_val = rd[1]
                qty_vals = rd[2:-1]
                total_val = rd[-1]
                row_fill = hour_color_map[h_val]

                col = start_col
                if h_val != prev_hour:
                    hc2 = ws.cell(row=current_row, column=col, value=h_val)
                    hc2.font = _font(bold=True, color=C_DARK_BLUE)
                    hc2.fill = _fill(C_PALE_BLUE)
                    hc2.alignment = _align()
                    hc2.border = _thin_border()
                else:
                    ec = ws.cell(row=current_row, column=col)
                    ec.fill = _fill(C_PALE_BLUE)
                    ec.border = _thin_border()
                    ec.alignment = _align()
                col += 1
                prev_hour = h_val

                wcc = ws.cell(row=current_row, column=col, value=wc_val)
                wcc.font = _font(bold=True, color=C_ORANGE_FONT)
                wcc.fill = _fill(C_ORANGE_FILL)
                wcc.alignment = _align()
                wcc.border = _thin_border()
                col += 1

                for q in qty_vals:
                    qc = ws.cell(row=current_row, column=col)
                    if q == 0:
                        qc.value = '—'
                        qc.font = _font(color=C_GREY_FONT)
                    else:
                        qc.value = q
                        qc.font = _font(color="FF000000")
                    qc.fill = _fill(row_fill)
                    qc.alignment = _align()
                    qc.border = _thin_border()
                    col += 1

                while col <= end_col - 1:
                    pc2 = ws.cell(row=current_row, column=col)
                    pc2.fill = _fill(row_fill)
                    pc2.border = _thin_border()
                    pc2.alignment = _align()
                    col += 1

                totc = ws.cell(row=current_row, column=end_col, value=total_val)
                totc.font = _font(bold=True, color="FF000000")
                totc.fill = _fill(C_YELLOW)
                totc.alignment = _align()
                totc.border = _thin_border()
                current_row += 1

            # TOTAL row for bay
            col = start_col
            tr_cell = ws.cell(row=current_row, column=col, value='TOTAL')
            tr_cell.font = _font(bold=True, color=C_WHITE)
            tr_cell.fill = _fill(C_GREEN)
            tr_cell.alignment = _align()
            tr_cell.border = _thin_border()
            col += 1

            wc_total = ws.cell(row=current_row, column=col)
            wc_total.fill = _fill(C_PALE_BLUE)
            wc_total.border = _thin_border()
            col += 1

            blk_totals = col_totals[2:-1]
            for bt in blk_totals:
                btc = ws.cell(row=current_row, column=col, value=bt)
                btc.font = _font(bold=True, color="FF000000")
                btc.fill = _fill(C_YELLOW)
                btc.alignment = _align()
                btc.border = _thin_border()
                col += 1

            while col <= end_col - 1:
                pc3 = ws.cell(row=current_row, column=col)
                pc3.fill = _fill(C_YELLOW)
                pc3.border = _thin_border()
                pc3.alignment = _align()
                col += 1

            gt_bay = col_totals[-1]
            gt_c = ws.cell(row=current_row, column=end_col, value=gt_bay)
            gt_c.font = _font(bold=True, color=C_WHITE)
            gt_c.fill = _fill(C_GREEN)
            gt_c.alignment = _align()
            gt_c.border = _thin_border()
            current_row += 1

        return current_row

    sts_next_rows = {sts: 2 for sts in sts_list}
    for sts in sts_list:
        next_row = write_detail_sts(
            ws_detail, sts,
            sts_col_start[sts],
            sts_table_widths[sts],
            detail_groups_by_sts[sts],
            sts_next_rows[sts]
        )
        sts_next_rows[sts] = next_row

    for c in range(1, total_detail_cols + 2):
        ws_detail.column_dimensions[get_column_letter(c)].width = 8

    # ============================================================
    # 8. Lưu file vào bộ nhớ đệm (BytesIO) thay vì ghi đĩa
    # ============================================================
    # MODIFIED: ghi workbook vào BytesIO thay vì file
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    # Tính các thông số thống kê để trả về
    total_rows = len(df_result)
    objective_value = pulp.value(prob.objective)   # giá trị hàm mục tiêu (tổng clash)

    # ADDED: trả về buffer và các thông số
    return output_buffer, total_rows, objective_value

# ============================================================
# Giữ khả năng chạy độc lập (khi file được thực thi trực tiếp)
# ============================================================
if __name__ == "__main__":
    # Ví dụ: chạy thử với file TEST2.xlsx (nếu có)
    try:
        buf, rows, obj = run_optimization('TEST2.xlsx')
        with open('optimized_allocation.xlsx', 'wb') as f:
            f.write(buf.read())
        print(f"Đã tạo file optimized_allocation.xlsx")
        print(f"Số dòng phân bổ: {rows}")
        print(f"Giá trị mục tiêu (tổng clash): {obj}")
    except Exception as e:
        print(f"Lỗi khi chạy thử: {e}")
