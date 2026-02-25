import streamlit as st
import pandas as pd
import io
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CẤU HÌNH TÊN SHEET DỮ LIỆU
# ==========================================
SHEET_MOVE  = 'MOVEHOUR-WEIGHTCLASS'
SHEET_BLOCK = 'BLOCK-WEIGHT CLASS'

# ==========================================
# MÀU SẮC DÙNG CHO EXCEL
# ==========================================
C_DARK   = '1F4E79'
C_MID    = '2E75B6'
C_LIGHT  = '9DC3E6'
C_VERYLI = 'D6E4F0'
C_ALT    = 'EBF3FB'
C_GREEN  = '375623'
C_YELLOW = 'FFF2CC'
C_RED_BG = 'FFE0E0'
C_RED_FG = 'C00000'
C_WHITE  = 'FFFFFF'
C_GREY   = 'BFBFBF'

# ==========================================
# CÁC HÀM XỬ LÝ DỮ LIỆU CHÍNH (Đã cập nhật thuật toán mới)
# ==========================================
def parse_input(input_file):
    xls = pd.ExcelFile(input_file)
    df1 = pd.read_excel(xls, sheet_name=SHEET_MOVE,  header=None)
    df2 = pd.read_excel(xls, sheet_name=SHEET_BLOCK, header=None)

    sts_row = df1.iloc[0].tolist()
    bay_row = df1.iloc[1].tolist()
    col_map = []   
    cur_sts = None
    for j, v in enumerate(sts_row):
        if pd.notna(v) and str(v).strip().startswith('STS'):
            cur_sts = str(v).strip()
        if j >= 2 and cur_sts and pd.notna(bay_row[j]) and str(bay_row[j]).strip():
            col_map.append((cur_sts, str(bay_row[j]).strip(), j))

    demand   = defaultdict(lambda: defaultdict(int))
    mh_order = []
    seen_mh  = set()
    cur_mh   = None

    for i in range(2, len(df1)):
        mh_raw = df1.iloc[i, 0]
        wc_v   = df1.iloc[i, 1]

        if pd.notna(mh_raw) and str(mh_raw).strip() not in ('', '0', 'nan'):
            cur_mh = str(mh_raw).strip()

        if cur_mh is None or not pd.notna(wc_v):
            continue
        try:
            wc = int(float(wc_v))
        except (ValueError, TypeError):
            continue

        if cur_mh not in seen_mh:
            mh_order.append(cur_mh)
            seen_mh.add(cur_mh)

        for sts, bay, pcol in col_map:
            v = df1.iloc[i, pcol]
            if pd.notna(v) and str(v).strip() not in ('', '0'):
                try:
                    demand[(cur_mh, sts, bay)][wc] += int(float(v))
                except (ValueError, TypeError):
                    pass

    blocks = []
    bwc    = {}
    for i in range(1, len(df2)):
        b = str(df2.iloc[i, 0]).strip()
        blocks.append(b)
        bwc[b] = {
            wc: (int(df2.iloc[i, wc]) if pd.notna(df2.iloc[i, wc]) else 0)
            for wc in [1, 2, 3, 4]
        }
    return col_map, demand, mh_order, blocks, bwc

def run_optimization(demand, mh_order, blocks, bwc):
    def score(wcs, b):
        """
        Block chỉ được chọn nếu có capacity > 0 cho TẤT CẢ WC có demand.
        Score = tổng block_qty tại các WC có demand.
        Trả về 0 nếu block không cover đủ tất cả WC -> không được chọn.
        """
        for wc, qty in wcs.items():
            if qty > 0 and bwc[b].get(wc, 0) == 0:
                return 0   # block không phù hợp với WC này
        return sum(bwc[b].get(wc, 0) for wc in wcs if wcs[wc] > 0)

    results = []
    for mh in mh_order:
        active = {(s, b): wcs for (m, s, b), wcs in demand.items() if m == mh and wcs}
        if not active: continue

        sorted_active = sorted(active.items(), key=lambda x: sum(1 for b in blocks if score(x[1], b) > 0))
        used   = set()
        assign = {}

        for (sts, bay), wcs in sorted_active:
            free_candidates = sorted([(score(wcs, b), b) for b in blocks if b not in used], reverse=True)
            if free_candidates and free_candidates[0][0] > 0:
                best_block = free_candidates[0][1]
                assign[(sts, bay)] = (best_block, False)   
                used.add(best_block)
            else:
                # Không còn block trống phù hợp → CLASH
                # Dùng partial score (bỏ qua ràng buộc all-WC) để chọn block ít tệ nhất
                partial = lambda b: sum(bwc[b].get(wc, 0) for wc in wcs if wcs[wc] > 0)
                all_candidates = sorted(
                    [(partial(b), b) for b in blocks],
                    reverse=True
                )
                best_block = all_candidates[0][1] if all_candidates else 'NO_MATCH'
                assign[(sts, bay)] = (best_block, True)    

        for (sts, bay), (blk, is_clash) in assign.items():
            for wc, qty in sorted(active[(sts, bay)].items()):
                results.append({
                    'mh'    : mh, 'sts'   : sts, 'bay'   : bay, 'wc'    : wc,
                    'qty'   : qty, 'blk'   : blk, 'clash' : 'CLASH' if is_clash else 'OK'
                })
    return pd.DataFrame(results)

# ==========================================
# CÁC HÀM XUẤT EXCEL
# ==========================================
def styled_cell(ws, row, col, value='', bold=False, color='000000', bg=None, size=10, align='center', wrap=False, border=True, font_name='Arial'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=font_name, bold=bold, color=color, size=size)
    if bg: c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if border:
        s = Side(style='thin', color='BFBFBF')
        c.border = Border(left=s, right=s, top=s, bottom=s)
    return c

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def write_result_sheet(wb, df_r):
    ws = wb.create_sheet('RESULT')
    headers = ['MOVE HOUR', 'STS', 'BAY', 'WEIGHT CLASS', 'CONTAINERS', 'ASSIGNED BLOCK', 'STATUS']
    widths  = [14, 8, 9, 13, 12, 16, 9]

    for j, (h, w) in enumerate(zip(headers, widths), 1):
        styled_cell(ws, 1, j, h, bold=True, color=C_WHITE, bg=C_DARK)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(df_r.itertuples(), 2):
        is_clash = row.clash == 'CLASH'
        bg  = C_RED_BG if is_clash else (C_ALT if i % 2 == 0 else C_WHITE)
        fg  = C_RED_FG if is_clash else '000000'
        for j, v in enumerate([row.mh, row.sts, row.bay, row.wc, row.qty, row.blk, row.clash], 1):
            styled_cell(ws, i, j, v, bold=is_clash, color=fg, bg=bg)
        ws.row_dimensions[i].height = 17

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{len(df_r) + 1}'

def write_matrix_sheet(wb, df_r, mh_order):
    ws = wb.create_sheet('MATRIX')
    df_agg = (df_r.groupby(['mh', 'sts', 'bay', 'blk'])['qty'].sum().reset_index())
    combos = (df_agg[['sts', 'bay', 'blk']].drop_duplicates().sort_values(['sts', 'bay', 'blk']).reset_index(drop=True))
    n = len(combos)

    lk = {}
    for _, r in df_agg.iterrows():
        for idx, c in combos.iterrows():
            if r.sts == c.sts and r.bay == c.bay and r.blk == c.blk:
                lk[(r.mh, idx)] = r.qty

    DC = 2
    DR = 6
    TOT_COL = DC + n

    merge(ws, 1, 1, 1, TOT_COL)
    styled_cell(ws, 1, 1, 'MA TRẬN PHÂN BỔ BLOCK  ▸  MOVE HOUR × STS / BAY / BLOCK', bold=True, color=C_DARK, bg=C_VERYLI, size=13, border=False)
    ws.row_dimensions[1].height = 30

    merge(ws, 2, 1, 4, 1)
    styled_cell(ws, 2, 1, 'MOVE\nHOUR', bold=True, color=C_WHITE, bg=C_DARK, wrap=True)
    ws.column_dimensions['A'].width = 13

    merge(ws, 2, TOT_COL, 4, TOT_COL)
    styled_cell(ws, 2, TOT_COL, 'TOTAL', bold=True, color=C_WHITE, bg=C_GREEN)
    ws.column_dimensions[get_column_letter(TOT_COL)].width = 9

    for idx, c in combos.iterrows():
        col = DC + idx
        styled_cell(ws, 2, col, c.sts,  bold=True, color=C_WHITE, bg=C_DARK,  size=9)
        styled_cell(ws, 3, col, c.bay,  bold=True, color=C_WHITE, bg=C_MID,   size=9)
        styled_cell(ws, 4, col, c.blk,  bold=True, color=C_DARK,  bg=C_LIGHT, size=9)
        ws.column_dimensions[get_column_letter(col)].width = 8

    for r in [2, 3, 4]:
        ws.row_dimensions[r].height = 20

    prev, start = combos.iloc[0].sts, DC
    for idx in range(1, n):
        if combos.iloc[idx].sts != prev:
            if DC + idx - 1 > start: merge(ws, 2, start, 2, DC + idx - 1)
            start = DC + idx
            prev  = combos.iloc[idx].sts
    merge(ws, 2, start, 2, DC + n - 1)

    prev_k = combos.iloc[0].sts + '|' + combos.iloc[0].bay
    start  = DC
    for idx in range(1, n):
        key = combos.iloc[idx].sts + '|' + combos.iloc[idx].bay
        if key != prev_k:
            if DC + idx - 1 > start: merge(ws, 3, start, 3, DC + idx - 1)
            start  = DC + idx
            prev_k = key
    merge(ws, 3, start, 3, DC + n - 1)

    ws.row_dimensions[5].height = 4

    for mi, mh in enumerate(mh_order):
        row    = DR + mi
        bg_row = C_ALT if mi % 2 == 0 else C_WHITE
        styled_cell(ws, row, 1, mh, bold=True, color=C_DARK, bg=C_VERYLI)
        ws.row_dimensions[row].height = 18
        row_total = 0

        for idx in range(n):
            col = DC + idx
            qty = lk.get((mh, idx), 0)
            if qty:
                styled_cell(ws, row, col, int(qty), bg=bg_row)
                row_total += qty
            else:
                c2 = ws.cell(row=row, column=col, value='—')
                c2.font      = Font(name='Arial', size=9, color=C_GREY)
                c2.fill      = PatternFill('solid', start_color=bg_row)
                c2.alignment = Alignment(horizontal='center', vertical='center')
                s = Side(style='thin', color='DCDCDC')
                c2.border    = Border(left=s, right=s, top=s, bottom=s)
        styled_cell(ws, row, TOT_COL, int(row_total), bold=True, bg=C_YELLOW)

    gt_row = DR + len(mh_order)
    styled_cell(ws, gt_row, 1, 'TOTAL', bold=True, color=C_WHITE, bg=C_GREEN)
    ws.row_dimensions[gt_row].height = 20
    grand = 0
    for idx in range(n):
        col     = DC + idx
        col_sum = int(sum(lk.get((mh, idx), 0) for mh in mh_order))
        if col_sum:
            styled_cell(ws, gt_row, col, col_sum, bold=True, bg=C_YELLOW)
        else:
            c2 = ws.cell(row=gt_row, column=col, value='—')
            c2.font      = Font(name='Arial', size=9, color=C_GREY, bold=True)
            c2.fill      = PatternFill('solid', start_color=C_YELLOW)
            c2.alignment = Alignment(horizontal='center', vertical='center')
        grand += col_sum
    styled_cell(ws, gt_row, TOT_COL, grand, bold=True, color=C_WHITE, bg=C_GREEN)
    ws.freeze_panes = ws.cell(row=DR, column=DC)

def write_summary_sheet(wb, df_r, mh_order):
    ws = wb.create_sheet('SUMMARY')
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    merge(ws, 1, 1, 1, 4)
    styled_cell(ws, 1, 1, 'KẾT QUẢ PHÂN BỔ BLOCK – TÓM TẮT', bold=True, color=C_DARK, bg=C_VERYLI, size=14, border=False)
    ws.row_dimensions[1].height = 30

    n_clash = int((df_r.clash == 'CLASH').sum())
    clash_color = C_RED_FG if n_clash else '006100'

    styled_cell(ws, 3, 1, 'Tổng số dòng phân bổ:', bold=True, border=False)
    styled_cell(ws, 3, 2, len(df_r), border=False)
    styled_cell(ws, 4, 1, 'Tổng CLASH:', bold=True, color=clash_color, border=False)
    styled_cell(ws, 4, 2, n_clash,      bold=True, color=clash_color, border=False)

    for j, h in enumerate(['TIME SLOT', 'STS/BAY ACTIVE', 'CLASH COUNT', 'TRẠNG THÁI'], 1):
        styled_cell(ws, 7, j, h, bold=True, color=C_WHITE, bg=C_DARK)
    
    for mi, mh in enumerate(mh_order):
        r   = 8 + mi
        sub = df_r[df_r.mh == mh]
        if sub.empty: continue
        n_act    = sub[['sts', 'bay']].drop_duplicates().shape[0]
        has_cl   = int((sub.clash == 'CLASH').sum() > 0)
        bg       = C_RED_BG if has_cl else (C_ALT if mi % 2 == 0 else C_WHITE)
        fg       = C_RED_FG if has_cl else '000000'
        status   = '⚠ CÓ CLASH' if has_cl else '✓ OK'
        for j, v in enumerate([mh, n_act, has_cl, status], 1):
            styled_cell(ws, r, j, v, bold=bool(has_cl), color=fg, bg=bg)

def write_detail_sheet_inline(wb, df_r_in=None, mh_order_in=None):
    df = df_r_in.copy()
    mh_all   = mh_order_in
    sts_list = sorted(df['sts'].unique())

    C = {'DARK': '1F4E79', 'MID': '2E75B6', 'LIGHT': 'BDD7EE', 'VERYLI': 'DEEAF1', 'ALT': 'EBF3FB', 'GREEN': '375623', 'YELLOW': 'FFF2CC', 'WHITE': 'FFFFFF', 'GREY': 'BFBFBF', 'WC_FG': '833C00', 'WC_BG': 'FCE4D6', 'MH_BG': 'D6E4F0'}
    t = Side(style='thin',   color='BFBFBF')
    def brd(): return Border(left=t, right=t, top=t, bottom=t)

    def sc(ws, r, c, val='', bold=False, fg='000000', bg=None, sz=10, align='center', wrap=False, italic=False, draw_border=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name='Arial', bold=bold, color=fg, size=sz, italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if bg: cell.fill = PatternFill('solid', start_color=bg)
        if draw_border: cell.border = brd()
        return cell

    def dash(ws, r, c, bg):
        cell = ws.cell(row=r, column=c, value='—')
        cell.font      = Font(name='Arial', size=9, color=C['GREY'])
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = brd()

    def mrgcell(ws, r1, c1, r2, c2, val='', bold=False, fg='000000', bg=None, sz=10, align='center'):
        if r2 > r1 or c2 > c1:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        sc(ws, r1, c1, val, bold, fg, bg, sz, align, draw_border=True)

    GAP = 1
    sts_info = {}
    for sts in sts_list:
        mh_rank = {mh: i for i, mh in enumerate(mh_all)}
        def bay_first_mh(bay):
            rows = df[(df['sts']==sts)&(df['bay']==bay)]
            if rows.empty: return 9999
            return mh_rank.get(rows['mh'].iloc[0], 9999)
        bays = sorted(df[df['sts']==sts]['bay'].unique(), key=bay_first_mh)
        max_blks = max(len(sorted(df[(df['sts']==sts)&(df['bay']==bay)]['blk'].unique())) for bay in bays)
        width = max_blks + 3
        sts_info[sts] = {'bays': bays, 'max_blks': max_blks, 'width': width}

    col_starts = {}
    cur_col = 1
    for sts in sts_list:
        col_starts[sts] = cur_col
        cur_col += sts_info[sts]['width'] + GAP
    total_width = cur_col - 1

    ws = wb.create_sheet('DETAIL')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_width)
    sc(ws, 1, 1, 'TỔNG HỢP CHI TIẾT  ▸  STS / BAY / MOVE HOUR / WC / BLOCK', bold=True, fg=C['DARK'], bg=C['VERYLI'], sz=13, draw_border=False)
    ws.row_dimensions[1].height = 28

    for sts in sts_list:
        cs = col_starts[sts]
        w  = sts_info[sts]['width']
        mrgcell(ws, 2, cs, 2, cs+w-1, sts, bold=True, fg=C['WHITE'], bg=C['DARK'], sz=12)
    ws.row_dimensions[2].height = 24

    START_ROW = 3

    for sts in sts_list:
        cs   = col_starts[sts]
        info = sts_info[sts]
        bays = info['bays']
        w    = info['width']
        cur_row = START_ROW

        for bay in bays:
            sub = df[(df['sts']==sts) & (df['bay']==bay)]
            blks  = sorted(sub['blk'].unique())
            n_blk = len(blks)

            c_mh    = cs 
            c_wc    = cs + 1
            c_blk0  = cs + 2 
            c_tot_fixed = cs + w - 1

            mrgcell(ws, cur_row, cs, cur_row, cs+w-1, bay, bold=True, fg=C['WHITE'], bg=C['MID'], sz=10)
            ws.row_dimensions[cur_row].height = 20
            cur_row += 1

            sc(ws, cur_row, c_mh, 'MOVE HOUR', bold=True, fg=C['WHITE'], bg=C['DARK'], sz=9)
            sc(ws, cur_row, c_wc, 'WC',         bold=True, fg=C['WHITE'], bg=C['DARK'], sz=9)
            for k, blk in enumerate(blks):
                sc(ws, cur_row, c_blk0+k, blk, bold=True, fg=C['WHITE'], bg=C['DARK'], sz=9)
            for k in range(n_blk, info['max_blks']):
                sc(ws, cur_row, c_blk0+k, '', bg=C['DARK'])
            sc(ws, cur_row, c_tot_fixed, 'TOTAL', bold=True, fg=C['WHITE'], bg=C['GREEN'], sz=9)
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

            col_tots   = defaultdict(int)
            grand_tot  = 0
            row_idx    = 0

            for mh in mh_all:
                mh_sub = sub[sub['mh'] == mh]
                if mh_sub.empty: continue
                wc_groups = sorted(mh_sub['wc'].unique())
                first_wc_in_mh = True

                for wc in wc_groups:
                    wc_sub = mh_sub[mh_sub['wc'] == wc]
                    bg = C['ALT'] if row_idx % 2 == 0 else C['WHITE']

                    if first_wc_in_mh:
                        sc(ws, cur_row, c_mh, mh, bold=True, fg=C['DARK'], bg=C['MH_BG'], sz=9)
                        first_wc_in_mh = False
                    else:
                        sc(ws, cur_row, c_mh, '', fg=C['DARK'], bg=C['MH_BG'], sz=9)

                    sc(ws, cur_row, c_wc, int(wc), bold=True, fg=C['WC_FG'], bg=C['WC_BG'], sz=9)

                    row_total = 0
                    for k, blk in enumerate(blks):
                        qty = int(wc_sub[wc_sub['blk']==blk]['qty'].sum())
                        if qty:
                            sc(ws, cur_row, c_blk0+k, qty, bg=bg, sz=10)
                            col_tots[blk] += qty
                            row_total += qty
                        else:
                            dash(ws, cur_row, c_blk0+k, bg)

                    for k in range(n_blk, info['max_blks']):
                        sc(ws, cur_row, c_blk0+k, '', bg=bg)

                    sc(ws, cur_row, c_tot_fixed, int(row_total), bold=True, bg=C['YELLOW'], sz=10)
                    grand_tot += row_total
                    ws.row_dimensions[cur_row].height = 17
                    row_idx += 1
                    cur_row += 1

            sc(ws, cur_row, c_mh, 'TOTAL', bold=True, fg=C['WHITE'], bg=C['GREEN'], sz=9)
            sc(ws, cur_row, c_wc, '',       bold=True, fg=C['WHITE'], bg=C['GREEN'], sz=9)
            for k, blk in enumerate(blks):
                v = int(col_tots[blk])
                sc(ws, cur_row, c_blk0+k, v, bold=True, bg=C['YELLOW'], sz=10)
            for k in range(n_blk, info['max_blks']):
                sc(ws, cur_row, c_blk0+k, '', bg=C['YELLOW'])
            sc(ws, cur_row, c_tot_fixed, int(grand_tot), bold=True, fg=C['WHITE'], bg=C['GREEN'], sz=10)
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1
            ws.row_dimensions[cur_row].height = 6
            cur_row += 1

    for sts in sts_list:
        cs = col_starts[sts]
        w  = sts_info[sts]['width']
        ws.column_dimensions[get_column_letter(cs)].width   = 11
        ws.column_dimensions[get_column_letter(cs+1)].width  = 5
        for k in range(sts_info[sts]['max_blks']):
            ws.column_dimensions[get_column_letter(cs+2+k)].width = 8
        ws.column_dimensions[get_column_letter(cs+w-1)].width = 8
        gap_col = cs + w
        if gap_col <= total_width:
            ws.column_dimensions[get_column_letter(gap_col)].width = 2
    ws.freeze_panes = 'A3'

def copy_source_sheets(wb, input_file):
    input_file.seek(0)
    wb_src = load_workbook(input_file)
    for sname in [SHEET_MOVE, SHEET_BLOCK]:
        ws_src = wb_src[sname]
        ws_dst = wb.create_sheet(sname)
        for row in ws_src.iter_rows():
            for c in row:
                ws_dst.cell(row=c.row, column=c.column, value=c.value)
        for col in ws_src.column_dimensions:
            w = ws_src.column_dimensions[col].width
            ws_dst.column_dimensions[col].width = w if w else 10

def build_excel_to_buffer(df_r, mh_order, input_file):
    wb = Workbook()
    wb.remove(wb.active)  

    copy_source_sheets(wb, input_file)      
    write_result_sheet(wb, df_r)     
    write_matrix_sheet(wb, df_r, mh_order)       
    write_summary_sheet(wb, df_r, mh_order) 
    write_detail_sheet_inline(wb, df_r, mh_order)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==========================================
# GIAO DIỆN WEB STREAMLIT
# ==========================================
st.set_page_config(page_title="Phân Bổ Block Optimizer", page_icon="🚢", layout="wide")

st.title("🚢 HỆ THỐNG TỐI ƯU PHÂN BỔ BLOCK TỰ ĐỘNG")
st.markdown("Hệ thống sẽ tự động đọc sheet `MOVEHOUR-WEIGHTCLASS` và `BLOCK-WEIGHT CLASS` từ file Excel của bạn, chạy thuật toán và xuất ra file kết quả được bôi màu chi tiết.")

uploaded_file = st.file_uploader("📂 Tải lên file Excel Input của bạn", type=["xlsx"])

if uploaded_file is not None:
    st.info("Đã nhận file. Bấm nút bên dưới để tiến hành tính toán.")
    
    if st.button("🚀 CHẠY THUẬT TOÁN TỐI ƯU", use_container_width=True):
        with st.spinner('Hệ thống đang xử lý và vẽ ma trận... Vui lòng đợi!'):
            try:
                col_map, demand, mh_order, blocks, bwc = parse_input(uploaded_file)
                df_result = run_optimization(demand, mh_order, blocks, bwc)
                excel_buffer = build_excel_to_buffer(df_result, mh_order, uploaded_file)
                
                st.success("✅ Tính toán hoàn tất!")
                n_clash = int((df_result.clash == 'CLASH').sum())
                
                # Hiển thị thống kê ngắn gọn
                col1, col2, col3 = st.columns(3)
                col1.metric("Tổng số lệnh phân bổ", f"{len(df_result)} dòng")
                col2.metric("Số lượng Time Slots", f"{len(mh_order)}")
                col3.metric("Số lượng Clash (Trùng lặp)", f"{n_clash} lỗi", delta_color="inverse")
                
                # Nút tải file
                st.download_button(
                    label="📥 TẢI FILE EXCEL KẾT QUẢ ĐẦY ĐỦ",
                    data=excel_buffer,
                    file_name="PHANBO_RESULT_FULL.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            except Exception as e:
                st.error(f"❌ Có lỗi xảy ra trong quá trình đọc file.\n\nChi tiết lỗi: {e}")
