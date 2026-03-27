import io
import pandas as pd
import numpy as np
import pulp
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from collections import defaultdict

# ============================================================
# COLOR PALETTE
# ============================================================
C_DARK_BLUE   = "FF1F4E79"
C_MID_BLUE    = "FF2E75B6"
C_LIGHT_BLUE  = "FF9DC3E6"
C_PALE_BLUE   = "FFD6E4F0"
C_ALT_ROW     = "FFEBF3FB"
C_WHITE       = "FFFFFFFF"
C_YELLOW      = "FFFFF2CC"
C_GREEN       = "FF375623"
C_TITLE_BG    = "FFDEEAF1"
C_TITLE_BG_M  = "FFD6E4F0"
C_ORANGE_FILL = "FFFCE4D6"
C_ORANGE_FONT = "FF833C00"
C_GREY_FONT   = "FFBFBFBF"
C_HEADER_BG   = "FFDEEAF1"

FONT_NAME = "Calibri"

def _font(bold=False, color="FF000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(border_style="medium", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _style(ws, coord, value=None, bold=False, font_color="FF000000",
           fill_color=None, align="center", wrap=False, border=True, size=10):
    cell = ws[coord] if isinstance(coord, str) else coord
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, color=font_color, size=size)
    if fill_color:
        cell.fill = _fill(fill_color)
    cell.alignment = _align(h=align, wrap=wrap)
    if border:
        cell.border = _thin_border()
    return cell

# ============================================================
# PUBLIC API
# ============================================================
def run_optimization(file_input):
    """
    Chạy thuật toán phân bổ tối ưu (phiên bản đã tối ưu tốc độ).

    Returns
    -------
    excel_buffer : io.BytesIO
    total_rows   : int
    total_clashes: int
    """
    # ============================================================
    # 1. Đọc dữ liệu đầu vào
    # ============================================================
    xls = pd.ExcelFile(file_input)

    # --- Sheet 1: MOVEHOUR-WEIGHTCLASS → demand ---
    df1 = pd.read_excel(xls, sheet_name='MOVEHOUR-WEIGHTCLASS', header=None)
    has_st_pod = (str(df1.iloc[1, 2]).strip().upper() == 'ST')
    data_col_start = 4 if has_st_pod else 2

    sts_bay_map = {}
    for col in range(data_col_start, df1.shape[1]):
        sts = df1.iloc[0, col]
        bay = df1.iloc[1, col]
        if pd.notna(sts) and pd.notna(bay):
            sts_bay_map[col] = (str(sts).strip(), str(bay).strip())

    demands = {}
    current_hour = None
    for idx in range(2, df1.shape[0]):
        row = df1.iloc[idx]
        hour = row[0]
        if pd.isna(hour):
            hour = current_hour
        else:
            current_hour = hour
        weight = row[1]
        if pd.isna(weight):
            continue
        weight = int(float(str(weight)))
        st_val  = str(row[2]).strip() if has_st_pod and pd.notna(row[2]) else ''
        pod_val = str(row[3]).strip() if has_st_pod and pd.notna(row[3]) else ''

        for col in range(data_col_start, df1.shape[1]):
            qty = row[col]
            if pd.notna(qty) and qty != '':
                qty = int(float(str(qty)))
                if qty > 0:
                    sts, bay = sts_bay_map[col]
                    key = (hour, sts, bay)
                    if key not in demands:
                        demands[key] = {}
                    dkey = (weight, st_val, pod_val)
                    demands[key][dkey] = demands[key].get(dkey, 0) + qty

    print(f"Demand format: {'WC+ST+POD' if has_st_pod else 'WC only (legacy)'}")
    job_keys = list(demands.keys())
    all_hours_sorted = sorted(set(h for (h, s, b) in job_keys))
    hour_rank = {h: i for i, h in enumerate(all_hours_sorted)}
    jobs_by_hour = {}
    for (h, s, b) in job_keys:
        jobs_by_hour.setdefault(h, []).append((s, b))

    # ── TĂNG TỐC #1: Pre-index jobs theo bay ──────────────────────
    # Tránh vòng lặp O(|job_keys|) lặp lại khi xây dựng constraints
    jobs_by_bay = defaultdict(list)
    for (h, s, bay) in job_keys:
        jobs_by_bay[bay].append((h, s, bay))

    # --- Sheet 2: BLOCK-WEIGHT CLASS → supply ---
    df2 = pd.read_excel(xls, sheet_name='BLOCK-WEIGHT CLASS', header=0)
    col_names = [str(c).strip() for c in df2.columns]
    has_st_pod_supply = (col_names[1].upper() == 'ST' and col_names[2].upper() == 'POD')
    wc_col_start = 3 if has_st_pod_supply else 1

    supply = {}
    blocks_set = set()
    for idx, row in df2.iterrows():
        block = str(row.iloc[0]).strip()
        if block in ('nan', 'GRAND TOTAL', '') or not block:
            continue
        st_v  = str(row.iloc[1]).strip() if has_st_pod_supply else ''
        pod_v = str(row.iloc[2]).strip() if has_st_pod_supply else ''
        skey  = (block, st_v, pod_v)
        wc_dict = {}
        for wi, w in enumerate([1, 2, 3, 4, 5]):
            col_idx = wc_col_start + wi
            if col_idx < len(row):
                val = row.iloc[col_idx]
                wc_dict[w] = int(val) if pd.notna(val) and val != '' else 0
            else:
                wc_dict[w] = 0
        supply[skey] = wc_dict
        blocks_set.add(block)

    weight_classes = [1, 2, 3, 4, 5]
    blocks = sorted(blocks_set)
    supply_keys = [k for k in supply if any(supply[k][w] > 0 for w in weight_classes)]
    print(f"Supply format: {'BLOCK+ST+POD' if has_st_pod_supply else 'BLOCK only (legacy)'}")
    print(f"Supply keys: {len(supply_keys)} (block×ST×POD combinations)")

    # --- Sheet 3: DATA (container-level) ---
    container_data_available = False
    try:
        df_containers = pd.read_excel(xls, sheet_name='DATA', header=0)
        cols = list(df_containers.columns)
        def find_col(candidates):
            for c in candidates:
                if c in cols: return c
            return None
        wc_src   = find_col(['YC', 'Unnamed: 1'])
        yp_src   = find_col(['YP', 'Unnamed: 2'])
        id_src   = find_col(['ID', 'Unnamed: 3'])
        st_src   = find_col(['ST'])
        pod_src  = find_col(['POD'])

        required_found = (wc_src and yp_src
                          and 'YB' in cols and 'YR' in cols and 'YT' in cols)
        if required_found:
            df_containers = df_containers.dropna(
                subset=[wc_src, yp_src, 'YB', 'YR', 'YT']
            ).copy()
            df_containers['REAL_WC']      = df_containers[wc_src].astype(float).astype(int)
            df_containers['YARD_POS']     = df_containers[yp_src].astype(str).str.strip()
            df_containers['REAL_CONT_ID'] = (df_containers[id_src].fillna('').astype(str).str.strip()
                                             if id_src else '')
            df_containers['CONT_ST']      = (df_containers[st_src].fillna('').astype(str).str.strip()
                                             if st_src else '')
            df_containers['CONT_POD']     = (df_containers[pod_src].fillna('').astype(str).str.strip()
                                             if pod_src else '')
            df_containers['YARD'] = df_containers['YARD'].astype(str).str.strip()
            df_containers['YB']   = df_containers['YB'].astype(float).astype(int)
            df_containers['YR']   = df_containers['YR'].astype(float).astype(int)
            df_containers['YT']   = df_containers['YT'].astype(float).astype(int)
            container_data_available = True
            print("Container-level DATA sheet found – stacking rules will be applied.")
            print(f"  {len(df_containers)} containers loaded.")
            print(f"  ST values : {sorted(df_containers['CONT_ST'].unique().tolist())}")
            print(f"  POD values: {sorted(df_containers['CONT_POD'].unique().tolist())}")
        else:
            print(f"DATA sheet missing required columns (need WC col + YP/YB/YR/YT). Skipped.")
    except Exception as e:
        print(f"No DATA sheet found – stacking rules skipped. ({e})")

    # --- Xây dựng cấu trúc stacking ---
    yb_wc_supply   = {}
    stack_ordering = {}
    blocking_pairs = []
    if container_data_available:
        df_c = df_containers[['YARD','YB','YR','YT','REAL_WC','YARD_POS','REAL_CONT_ID','CONT_ST','CONT_POD']].copy()
        for block in blocks:
            block_df = df_c[df_c['YARD'] == block]
            if block_df.empty:
                continue
            yb_wc_supply[block] = {}
            stack_ordering[block] = {}
            for yb, yb_df in block_df.groupby('YB'):
                yb_wc_supply[block][yb] = {}
                stack_ordering[block][yb] = {}
                for wc, cnt in yb_df.groupby('REAL_WC').size().items():
                    yb_wc_supply[block][yb][wc] = int(cnt)
                for yr, yr_df in yb_df.groupby('YR'):
                    ordered = yr_df.sort_values('YT', ascending=False)[['YT','REAL_WC']].values.tolist()
                    stack_ordering[block][yb][yr] = [(int(t), int(w)) for t, w in ordered]
                for yr, tiers in stack_ordering[block][yb].items():
                    wcs_above = []
                    for tier, wc in tiers:
                        if wcs_above:
                            for (prev_wc, prev_tier) in wcs_above:
                                if prev_wc != wc:
                                    blocking_pairs.append((block, yb, yr, prev_tier, prev_wc, tier, wc))
                        wcs_above.append((wc, tier))
        print(f"Stacking structures built: {len(blocking_pairs)} cross-WC blocking pairs found.")

    # ============================================================
    # 2. Kiểm tra cân bằng demand – supply
    # ============================================================
    total_demand = {}
    for job in job_keys:
        for dkey, qty in demands[job].items():
            total_demand[dkey] = total_demand.get(dkey, 0) + qty

    total_supply = {}
    for skey in supply_keys:
        block, st_v, pod_v = skey
        for w in weight_classes:
            tkey = (w, st_v, pod_v)
            total_supply[tkey] = total_supply.get(tkey, 0) + supply[skey][w]

    print("Total demand per (WC, ST, POD):")
    for k in sorted(total_demand):
        print(f"  WC={k[0]} ST={k[1]} POD={k[2]}: {total_demand[k]}")
    print("Total supply per (WC, ST, POD):")
    for k in sorted(total_supply):
        print(f"  WC={k[0]} ST={k[1]} POD={k[2]}: {total_supply[k]}")

    ok = True
    for k in set(list(total_demand.keys()) + list(total_supply.keys())):
        d = total_demand.get(k, 0)
        s = total_supply.get(k, 0)
        if d != s:
            print(f"ERROR Mismatch WC={k[0]} ST={k[1]} POD={k[2]}: demand={d}, supply={s}")
            ok = False
    if not ok:
        raise ValueError("Demand/supply mismatch — kiểm tra lại file input.")
    print("Demand/supply balanced OK.")

    # ============================================================
    # 3. Xây dựng và giải mô hình tối ưu
    # ============================================================
    prob = pulp.LpProblem("Minimize_Clashes_ST_POD", pulp.LpMinimize)

    y_vars = {}
    for (h, s, bay) in job_keys:
        for b in blocks:
            y_vars[(h, s, bay, b)] = pulp.LpVariable(f"y_{h}_{s}_{bay}_{b}", cat='Binary')

    x_vars = {}
    for (h, s, bay) in job_keys:
        for dkey in demands[(h, s, bay)]:
            w, st_v, pod_v = dkey
            for skey in supply_keys:
                b, sup_st, sup_pod = skey
                if sup_st != st_v or sup_pod != pod_v:
                    continue
                vname = f"x_{h}_{s}_{bay}_{b}_{w}_{st_v}_{pod_v}"
                x_vars[(h, s, bay, b, dkey)] = pulp.LpVariable(vname, lowBound=0, cat='Integer')

    u_vars = {}
    e_vars = {}
    for h in jobs_by_hour:
        for b in blocks:
            u_vars[(h, b)] = pulp.LpVariable(f"u_{h}_{b}", lowBound=0, cat='Integer')
            e_vars[(h, b)] = pulp.LpVariable(f"e_{h}_{b}", lowBound=0, cat='Integer')
            prob += u_vars[(h, b)] == pulp.lpSum(y_vars[(h, s, bay, b)] for (s, bay) in jobs_by_hour[h])
            prob += e_vars[(h, b)] >= u_vars[(h, b)] - 1

    CLASH_W        = 100.0
    SINGLE_W       = 10.0
    SPREAD_W       = 5.0
    BLOCK_BAY_WC_W = 2.0
    BAY_SINGLE_W   = 10.0

    single_block = {}
    for (h, s, bay) in job_keys:
        single_block[(h, s, bay)] = pulp.LpVariable(f"sb_{h}_{s}_{bay}", lowBound=0, upBound=1, cat='Continuous')
        prob += single_block[(h, s, bay)] >= (2 - pulp.lpSum(y_vars[(h, s, bay, b)] for b in blocks))

    all_bays = sorted(set(bay for (_, _, bay) in job_keys))

    # ── TĂNG TỐC #2: Pre-index x_vars theo (block, bay, wc) ──────
    # Tránh triple nested loop O(B×BA×W×J×D) → O(1) lookup
    x_by_block_bay_wc = defaultdict(list)
    for (h, s, bay, b, dkey), xvar in x_vars.items():
        w = dkey[0]
        x_by_block_bay_wc[(b, bay, w)].append((xvar, demands[(h, s, bay)][dkey]))

    block_bay = {}
    for b in blocks:
        for bay in all_bays:
            var = pulp.LpVariable(f"bb_{b}_{bay}", cat='Binary')
            block_bay[(b, bay)] = var
            # ── TĂNG TỐC #2a: dùng jobs_by_bay thay vì duyệt toàn bộ job_keys
            for (h, s, bj) in jobs_by_bay[bay]:
                prob += var >= y_vars[(h, s, bay, b)]

    block_bay_wc = {}
    for b in blocks:
        for bay in all_bays:
            for wc in weight_classes:
                entries = x_by_block_bay_wc.get((b, bay, wc), [])
                if not entries:
                    continue  # bỏ qua biến không cần thiết → giảm kích thước model
                var = pulp.LpVariable(f"bbw_{b}_{bay}_{wc}", cat='Binary')
                block_bay_wc[(b, bay, wc)] = var
                # Mỗi entry đã được index sẵn, không cần duyệt lại
                for xvar, d in entries:
                    prob += var >= xvar / (d + 0.1)

    bay_single = {}
    for bay in all_bays:
        var = pulp.LpVariable(f"bs_{bay}", lowBound=0, upBound=1, cat='Continuous')
        bay_single[bay] = var
        total_blocks_bay = pulp.lpSum(block_bay[(b, bay)] for b in blocks)
        prob += var >= (2 - total_blocks_bay)

    min_blocks_per_bay = 2
    for bay in all_bays:
        prob += pulp.lpSum(block_bay[(b, bay)] for b in blocks) >= min_blocks_per_bay

    clash_term        = pulp.lpSum(e_vars.values())
    single_term       = pulp.lpSum(single_block.values())
    spread_term       = pulp.lpSum(block_bay.values())
    block_bay_wc_term = pulp.lpSum(block_bay_wc.values())
    bay_single_term   = pulp.lpSum(bay_single.values())

    prob += (CLASH_W * clash_term +
             SINGLE_W * single_term +
             SPREAD_W * spread_term +
             BLOCK_BAY_WC_W * block_bay_wc_term +
             BAY_SINGLE_W * bay_single_term)

    # Ràng buộc demand
    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            w, st_v, pod_v = dkey
            x_sum = pulp.lpSum(x_vars[(h, s, bay, b, dkey)]
                               for skey in supply_keys
                               for b in [skey[0]]
                               if skey[1] == st_v and skey[2] == pod_v
                               and (h, s, bay, b, dkey) in x_vars)
            prob += x_sum == d

    for skey in supply_keys:
        b, st_v, pod_v = skey
        for w in weight_classes:
            dkey_w = [(h, s, bay, (w, st_v, pod_v))
                      for (h, s, bay) in job_keys
                      if (w, st_v, pod_v) in demands[(h, s, bay)]]
            if not dkey_w:
                continue
            prob += pulp.lpSum(x_vars[(h, s, bay, b, (w, st_v, pod_v))]
                               for (h, s, bay, dk) in dkey_w
                               if (h, s, bay, b, dk) in x_vars) <= supply[skey][w]

    for (h, s, bay) in job_keys:
        for dkey, d in demands[(h, s, bay)].items():
            for skey in supply_keys:
                b = skey[0]
                if (h, s, bay, b, dkey) in x_vars:
                    prob += x_vars[(h, s, bay, b, dkey)] <= d * y_vars[(h, s, bay, b)]

    # ── TĂNG TỐC #3: Thử HiGHS trước (nhanh hơn CBC ~3-5x), fallback CBC ──
    solver = _get_best_solver(time_limit=300)
    prob.solve(solver)

    status = prob.status
    print(f"Status: {pulp.LpStatus[status]}")
    if status == pulp.LpStatusInfeasible:
        raise RuntimeError("Model infeasible — kiểm tra supply/demand và ràng buộc.")
    elif status not in (1,):
        print("No optimal solution found within time limit – using best solution found.")

    # ============================================================
    # 4. Trích xuất kết quả và gán container
    # ============================================================
    result_rows = []
    for (h, s, bay, b) in y_vars:
        if pulp.value(y_vars[(h, s, bay, b)]) is not None and pulp.value(y_vars[(h, s, bay, b)]) > 0.5:
            for dkey in demands[(h, s, bay)]:
                w, st_v, pod_v = dkey
                xkey = (h, s, bay, b, dkey)
                if xkey not in x_vars:
                    continue
                qty = pulp.value(x_vars[xkey])
                if qty is not None and qty > 0.5:
                    result_rows.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay,
                        'ASSIGNED BLOCK': b,
                        'WEIGHT CLASS': w, 'ST': st_v, 'POD': pod_v,
                        'QUANTITIES': int(round(qty))
                    })

    df_result = pd.DataFrame(result_rows)
    df_result.sort_values(['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK'], inplace=True)

    df_result_detail = []
    if container_data_available:
        # ── TĂNG TỐC #4: Build pool với dict lookup nhanh ────────────
        pool = {}
        for _, row in df_containers[['YARD','YB','YR','YT','REAL_WC',
                                     'YARD_POS','REAL_CONT_ID',
                                     'CONT_ST','CONT_POD']].iterrows():
            blk = row['YARD']
            pool.setdefault(blk, []).append({
                'yb': int(row['YB']), 'yr': int(row['YR']), 'yt': int(row['YT']),
                'wc': int(row['REAL_WC']),
                'yard_pos':     row['YARD_POS'],
                'real_cont_id': row['REAL_CONT_ID'],
                'st':           row['CONT_ST'],
                'pod':          row['CONT_POD'],
                'picked': False, 'pick_h': None
            })

        # ── TĂNG TỐC #5: Pre-compute blockers (containers nằm trên) ──
        # Thay vì duyệt toàn bộ pool trong accessible_at() mỗi lần,
        # ta tính 1 lần duy nhất: mỗi container → list container chặn nó
        blockers_map = {}   # id(cont) -> list[cont dict]
        for blk, conts in pool.items():
            # Group by (yb, yr) để tìm blockers nhanh
            stack_map = defaultdict(list)
            for c in conts:
                stack_map[(c['yb'], c['yr'])].append(c)
            for c in conts:
                above = [other for other in stack_map[(c['yb'], c['yr'])]
                         if other is not c and other['yt'] > c['yt']]
                blockers_map[id(c)] = above

        opened_ybs = set()

        def accessible_at(cont, h_rank_val):
            # O(blockers) thay vì O(tất cả containers trong block)
            for blocker in blockers_map.get(id(cont), []):
                if not blocker['picked']:
                    return False
                if blocker['pick_h'] is not None and hour_rank[blocker['pick_h']] > h_rank_val:
                    return False
            return True

        def pick_n(block, wc, st_match, pod_match, qty, h, s_job, bay_job, h_rank_val, result_list):
            containers = pool[block]
            remaining  = qty

            def matches(c):
                if c['wc'] != wc:                              return False
                if st_match  and c.get('st','')  != st_match: return False
                if pod_match and c.get('pod','') != pod_match: return False
                return True

            while remaining > 0:
                cands = [c for c in containers
                         if not c['picked'] and matches(c)
                         and accessible_at(c, h_rank_val)]
                if not cands:
                    break

                # ── TĂNG TỐC #6: tính yb_cnt dùng Counter thay vì dict.get ──
                yb_cnt = defaultdict(int)
                for c in cands:
                    yb_cnt[c['yb']] += 1

                cands.sort(key=lambda c: (
                    0 if (block, c['yb']) in opened_ybs else 1,
                    -yb_cnt[c['yb']],
                    c['yb'],
                    c['yr'],
                    -c['yt']
                ))
                best = cands[0]
                best['picked'] = True
                best['pick_h'] = h
                opened_ybs.add((block, best['yb']))
                result_list.append({
                    'MOVE HOUR':      h,
                    'CONTAINER ID':   best['real_cont_id'],
                    'ST':             best.get('st', st_match),
                    'POD':            best.get('pod', pod_match),
                    'STS': s_job,     'BAY': bay_job,
                    'ASSIGNED BLOCK': block,
                    'WEIGHT CLASS':   wc,
                    'QUANTITIES':     qty,
                    'YB': best['yb'], 'YR': best['yr'], 'YT': best['yt'],
                    'YARD POSITION':  best['yard_pos']
                })
                remaining -= 1
            return remaining

        df_result_sorted = df_result.copy()
        df_result_sorted['_hr'] = df_result_sorted['MOVE HOUR'].map(hour_rank)
        df_result_sorted.sort_values(['_hr','STS','BAY','ASSIGNED BLOCK','WEIGHT CLASS'],
                                      inplace=True)

        deferred = []
        for h in all_hours_sorted:
            h_rank_val = hour_rank[h]
            hour_asgns = df_result_sorted[df_result_sorted['MOVE HOUR'] == h]
            for _, asg in hour_asgns.iterrows():
                s, bay_job, b = asg['STS'], asg['BAY'], asg['ASSIGNED BLOCK']
                w    = int(asg['WEIGHT CLASS'])
                st_v = str(asg.get('ST', '')).strip()
                pod_v= str(asg.get('POD', '')).strip()
                qty  = int(asg['QUANTITIES'])
                if b not in pool:
                    df_result_detail.append({
                        'MOVE HOUR': h, 'STS': s, 'BAY': bay_job,
                        'ASSIGNED BLOCK': b, 'WEIGHT CLASS': w,
                        'CONTAINER ID': '', 'ST': '', 'POD': '', 'QUANTITIES': qty,
                        'YB': '', 'YR': '', 'YT': '', 'YARD POSITION': ''
                    })
                    continue
                rem = pick_n(b, w, st_v, pod_v, qty, h, s, bay_job, h_rank_val, df_result_detail)
                if rem > 0:
                    deferred.append({'b': b, 'wc': w, 'st': st_v, 'pod': pod_v,
                                      'qty': rem, 'h_orig': h, 's': s, 'bay': bay_job,
                                      'h_rank_min': h_rank_val})
            still_deferred = []
            for d in deferred:
                rem = pick_n(d['b'], d['wc'], d.get('st',''), d.get('pod',''),
                             d['qty'], h, d['s'], d['bay'], h_rank_val, df_result_detail)
                if rem > 0:
                    d2 = d.copy()
                    d2['qty'] = rem
                    still_deferred.append(d2)
            deferred = still_deferred

        rehandle_count = sum(d['qty'] for d in deferred)
        if rehandle_count > 0:
            print(f"  INFO: {rehandle_count} containers require re-handling (cross-WC stacking unavoidable):")
            for d in deferred:
                print(f"    Block {d['b']} WC{d['wc']} x{d['qty']} (originally at {d['h_orig']})")
        else:
            print("  All containers assigned with no re-handling required.")

        df_result_detail = pd.DataFrame(df_result_detail)

    else:
        df_result_detail = df_result.copy()
        df_result_detail.insert(1, 'CONTAINER ID', '')
        df_result_detail.insert(2, 'ST', '')
        df_result_detail.insert(3, 'POD', '')
        df_result_detail['YB'] = ''
        df_result_detail['YR'] = ''
        df_result_detail['YT'] = ''
        df_result_detail['YARD POSITION'] = ''

    df_result_detail.sort_values(
        ['MOVE HOUR', 'STS', 'BAY', 'ASSIGNED BLOCK', 'WEIGHT CLASS', 'YB', 'YR', 'YT'],
        inplace=True
    )

    # ============================================================
    # 5. Trích xuất clash
    # ============================================================
    clash_details = []
    total_clashes = 0
    for (h, b) in e_vars:
        e_val = pulp.value(e_vars[(h, b)])
        if e_val is not None and e_val > 0.5:
            total_clashes += e_val
            u_val = pulp.value(u_vars[(h, b)])
            jobs = []
            for (s, bay) in jobs_by_hour.get(h, []):
                y_key = (h, s, bay, b)
                if y_key in y_vars and pulp.value(y_vars[y_key]) > 0.5:
                    jobs.append(f"{s}@{bay}")
            clash_details.append({
                'MOVE HOUR': h,
                'BLOCK': b,
                'SỐ LƯỢNG BAY (u)': int(u_val) if u_val is not None else 0,
                'CLASH (e = u-1)': int(e_val),
                'DANH SÁCH JOB (STS@BAY)': ', '.join(jobs)
            })
    df_clash = pd.DataFrame(clash_details)
    if not df_clash.empty:
        df_clash.sort_values(['MOVE HOUR', 'BLOCK'], inplace=True)
    print(f"Total clashes (e sum): {total_clashes}")

    # ============================================================
    # 6. Ghi file Excel
    # ============================================================
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ----- Sheet CLASH -----
    ws_clash = wb.create_sheet('CLASH')
    headers_clash = ['MOVE HOUR', 'BLOCK', 'SỐ LƯỢNG BAY (u)', 'CLASH (e = u-1)', 'DANH SÁCH JOB (STS@BAY)']
    for c_idx, hdr in enumerate(headers_clash, 1):
        cell = ws_clash.cell(row=1, column=c_idx, value=hdr)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK_BLUE)
        cell.alignment = _align()
        cell.border = _thin_border()

    if not df_clash.empty:
        for r_idx, row in enumerate(df_clash.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws_clash.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _font()
                cell.fill = _fill(C_WHITE if r_idx % 2 == 0 else C_ALT_ROW)
                cell.alignment = _align()
                cell.border = _thin_border()
    else:
        cell = ws_clash.cell(row=2, column=1, value='Không có clash nào xảy ra.')
        cell.font = _font()
        cell.fill = _fill(C_WHITE)
        cell.alignment = _align()
        ws_clash.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    ws_clash.column_dimensions['A'].width = 14
    ws_clash.column_dimensions['B'].width = 12
    ws_clash.column_dimensions['C'].width = 18
    ws_clash.column_dimensions['D'].width = 18
    ws_clash.column_dimensions['E'].width = 50

    # ----- Định nghĩa cột cho các sheet RESULT -----
    core_cols     = ['MOVE HOUR', 'CONT LIST', 'CONTAINER ID',
                      'ST', 'POD', 'STS', 'BAY',
                      'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']
    position_cols = ['YB', 'YR', 'YT', 'YARD POSITION']
    if container_data_available:
        all_result_cols = core_cols + position_cols
    else:
        all_result_cols = ['MOVE HOUR', 'STS', 'BAY',
                            'ASSIGNED BLOCK', 'WEIGHT CLASS', 'QUANTITIES']

    CONT_LIST_COLS = {'CONT LIST'}
    CONT_ID_COLS   = {'CONTAINER ID', 'ST', 'POD'}
    POSITION_COLS  = set(position_cols)

    col_widths = {
        'MOVE HOUR': 14, 'CONT LIST': 45, 'CONTAINER ID': 20,
        'ST': 10,        'POD': 10,       'STS': 10,  'BAY': 10,
        'ASSIGNED BLOCK': 16, 'WEIGHT CLASS': 14, 'QUANTITIES': 12,
        'YB': 8, 'YR': 8, 'YT': 8, 'YARD POSITION': 18,
    }

    def write_result_sheet(ws, df, sheet_title):
        n_rows = len(df)
        cont_list_map = {}
        if container_data_available and 'CONTAINER ID' in df.columns:
            for (mh, bay), grp in df.groupby(['MOVE HOUR', 'BAY']):
                ids = [str(v).strip() for v in grp['CONTAINER ID']
                       if str(v).strip() not in ('', 'nan')]
                cont_list_map[(mh, bay)] = ', '.join(ids) if ids else ''

        # Header
        for c_idx, cn in enumerate(all_result_cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=cn)
            if cn in CONT_LIST_COLS:
                cell.fill = _fill(C_PALE_BLUE)
            elif cn in CONT_ID_COLS:
                cell.fill = _fill(C_MID_BLUE)
            elif cn in POSITION_COLS:
                cell.fill = _fill(C_LIGHT_BLUE)
            else:
                cell.fill = _fill(C_DARK_BLUE)
            cell.font      = _font(bold=True, color=C_WHITE)
            cell.alignment = _align(wrap=True)
            cell.border    = _thin_border()

        merge_groups = []
        cont_list_col_idx = all_result_cols.index('CONT LIST') + 1 if 'CONT LIST' in all_result_cols else None
        if container_data_available and cont_list_col_idx:
            prev_key  = None
            grp_start = 2
            for i, (_, row) in enumerate(df.iterrows()):
                cur_key   = (row.get('MOVE HOUR', ''), row.get('BAY', ''))
                excel_row = i + 2
                if cur_key != prev_key:
                    if prev_key is not None:
                        merge_groups.append((prev_key, grp_start, excel_row - 1,
                                             cont_list_map.get(prev_key, '')))
                    prev_key  = cur_key
                    grp_start = excel_row
            if prev_key is not None:
                merge_groups.append((prev_key, grp_start, n_rows + 1,
                                     cont_list_map.get(prev_key, '')))

        # ── TĂNG TỐC #7: Ghi data rows bằng bulk append thay vì cell-by-cell ──
        group_key   = None
        group_shade = C_ALT_ROW
        fill_white   = _fill(C_WHITE)
        fill_alt     = _fill(C_ALT_ROW)
        border       = _thin_border()
        align_center = _align()
        font_default = _font(color='FF000000')

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            this_key = (row.get('MOVE HOUR'), row.get('STS'), row.get('BAY'),
                        row.get('ASSIGNED BLOCK'), row.get('WEIGHT CLASS'))
            if this_key != group_key:
                group_shade = C_WHITE if group_shade == C_ALT_ROW else C_ALT_ROW
                group_key = this_key
            current_fill = fill_white if group_shade == C_WHITE else fill_alt

            for c_idx, cn in enumerate(all_result_cols, 1):
                if cn == 'CONT LIST':
                    val = None
                else:
                    val = row.get(cn, '')
                    if cn in ('YB', 'YR', 'YT') and val != '':
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            pass
                    if val == '' or (isinstance(val, float) and str(val) == 'nan'):
                        val = None

                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = font_default
                cell.fill      = current_fill
                cell.alignment = align_center
                cell.border    = border

        if cont_list_col_idx:
            align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
            fill_pale      = _fill(C_PALE_BLUE)
            font_small     = _font(color='FF000000', size=9)
            for (mh, bay), r_start, r_end, list_text in merge_groups:
                cell = ws.cell(row=r_start, column=cont_list_col_idx,
                               value=list_text or None)
                cell.font      = font_small
                cell.fill      = fill_pale
                cell.alignment = align_top_left
                cell.border    = border
                if r_end > r_start:
                    ws.merge_cells(
                        start_row=r_start, start_column=cont_list_col_idx,
                        end_row=r_end,     end_column=cont_list_col_idx
                    )
                    ws.cell(row=r_start, column=cont_list_col_idx).alignment = align_top_left

            for (mh, bay), r_start, r_end, list_text in merge_groups:
                span     = r_end - r_start + 1
                n_ids    = len([x for x in list_text.split(',') if x.strip()]) if list_text else 0
                rows_needed = max(1, -(-n_ids // max(1, span)))
                rh = max(15, min(60, rows_needed * 13))
                for r in range(r_start, r_end + 1):
                    ws.row_dimensions[r].height = rh

        for c_idx, cn in enumerate(all_result_cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(cn, 14)

        print(f"  Sheet '{sheet_title}': {n_rows} rows written.")

    # ----- Tạo các sheet RESULT theo ST -----
    if container_data_available and 'ST' in df_result_detail.columns:
        st_values = sorted(df_result_detail['ST'].dropna().unique().tolist())
        st_values = [s for s in st_values if str(s).strip() not in ('', 'nan')]
    else:
        st_values = ['ALL']
    if not st_values:
        st_values = ['ALL']

    for st_idx, st_val in enumerate(st_values, 1):
        sheet_name = f"RESULT {st_idx} ({st_val})" if st_val != 'ALL' else 'RESULT'
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)
        if st_val == 'ALL':
            df_rd = df_result_detail.reset_index(drop=True)
        else:
            df_rd = df_result_detail[df_result_detail['ST'].astype(str).str.strip() == str(st_val).strip()].reset_index(drop=True)
        write_result_sheet(ws, df_rd, sheet_name)

    # ----- Sheet RESULT TOTAL -----
    ws_total = wb.create_sheet('RESULT TOTAL')
    write_result_sheet(ws_total, df_result_detail.reset_index(drop=True), 'RESULT TOTAL')

    # ============================================================
    # 7. Lưu vào buffer
    # ============================================================
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    total_rows = len(df_result_detail)
    print(f"Done. Rows={total_rows}, Total Clashes={total_clashes}")
    return excel_buffer, total_rows, total_clashes


# ============================================================
# HELPER: Chọn solver tốt nhất khả dụng
# ── TĂNG TỐC #3: HiGHS nhanh hơn CBC ~3-5x cho MIP ──────────
# ============================================================
def _get_best_solver(time_limit=300):
    """
    Thử theo thứ tự ưu tiên:
      1. HiGHS  – solver MIP hiện đại, nhanh hơn CBC đáng kể
      2. GLPK   – fallback nếu HiGHS không cài
      3. CBC    – mặc định của PuLP (luôn có sẵn)
    """
    # --- HiGHS ---
    try:
        highs = pulp.HiGHS_CMD(
            msg=True,
            timeLimit=time_limit,
            options=[
                ("parallel", "on"),           # dùng nhiều CPU
                ("threads",  str(_cpu_count())),
            ]
        )
        # Kiểm tra HiGHS có thực sự cài không
        test = pulp.LpProblem("_test", pulp.LpMinimize)
        x = pulp.LpVariable("x")
        test += x
        test += x >= 0
        test.solve(highs)
        print(f"[Solver] Sử dụng HiGHS ({_cpu_count()} threads)")
        return highs
    except Exception:
        pass

    # --- CBC với multi-thread ---
    try:
        n = _cpu_count()
        cbc = pulp.PULP_CBC_CMD(msg=True, timeLimit=time_limit, threads=n)
        print(f"[Solver] Sử dụng CBC ({n} threads)")
        return cbc
    except Exception:
        pass

    # --- CBC fallback ---
    print("[Solver] Sử dụng CBC (single thread – fallback)")
    return pulp.PULP_CBC_CMD(msg=True, timeLimit=time_limit)


def _cpu_count():
    import os
    try:
        return max(1, os.cpu_count() or 1)
    except Exception:
        return 1
